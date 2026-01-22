import openpyxl
from datetime import datetime
import os
import re


def format_sql_value(value):
    """
    格式化SQL值
    
    Args:
        value: 原始值
        
    Returns:
        格式化后的SQL值
    """
    if value is None or value == '':
        return 'NULL'
    elif isinstance(value, str):
        escaped_value = value.replace("'", "\\'")
        return f"'{escaped_value}'"
    elif isinstance(value, (int, float)):
        return str(value)
    else:
        return f"'{str(value)}'"


def extract_fin_spacing_num(fin_spacing):
    """
    从fin_spacing字段中提取数字（支持整数和小数）
    例如：从 "C04=4mm" 中提取 4.0，从 "C04=4.5mm" 中提取 4.5
    
    Args:
        fin_spacing: fin_spacing字段的原始值
        
    Returns:
        提取的float数字，如果无法提取则返回None
    """
    if fin_spacing is None or fin_spacing == '':
        return None
    
    # 使用正则表达式匹配等号后的数字（支持整数和小数，例如：C04=4mm 或 C04=4.5mm）
    match = re.search(r'=(\d+\.?\d*)', str(fin_spacing))
    if match:
        try:
            # 转换为float类型，确保SQL中输出为数字值而不是字符串
            return float(match.group(1))
        except ValueError:
            return None
    
    return None


def generate_cooling_capacity_insert(sheet, col_idx, refrigerant):
    """
    生成cooling_capacity表的INSERT语句
    
    Args:
        sheet: Excel工作簿对象
        col_idx: 列索引（从1开始）
        refrigerant: 制冷剂
        
    Returns:
        INSERT语句列表
    """
    cooler_id = sheet.cell(1, col_idx).value
    
    if cooler_id is None or cooler_id == '':
        return []
    
    insert_statements = []
    
    working_status_map = {
        2: 'SC1',
        3: 'SC2',
        4: 'SC3',
        5: 'SC4',
        6: 'SC5'
    }
    
    for row_idx, working_status in working_status_map.items():
        capacity = sheet.cell(row_idx, col_idx).value
        
        if capacity is None or capacity == '':
            continue
        
        cooler_id_sql = format_sql_value(cooler_id)
        working_status_sql = format_sql_value(working_status)
        refrigerant_sql = format_sql_value(refrigerant)
        capacity_sql = format_sql_value(capacity)
        
        insert_statement = (
            f"INSERT INTO cooling_capacity (cooler_id, working_status, refrigerant, capacity, "
            f"created_time, updated_time, is_deleted) "
            f"VALUES ({cooler_id_sql}, {working_status_sql}, {refrigerant_sql}, {capacity_sql}, "
            f"NOW(), NOW(), 0);"
        )
        insert_statements.append(insert_statement)
    
    return insert_statements


def generate_cooler_insert(sheet, col_idx):
    """
    生成cooler表的INSERT语句
    
    Args:
        sheet: Excel工作簿对象
        col_idx: 列索引（从1开始）
        
    Returns:
        INSERT语句
    """
    model = sheet.cell(1, col_idx).value
    
    if model is None or model == '':
        return None
    
    heat_exchange_area = sheet.cell(8, col_idx).value
    tube_volumn = sheet.cell(9, col_idx).value
    air_flow_rate = sheet.cell(10, col_idx).value
    total_fan_power = sheet.cell(11, col_idx).value
    total_fan_current = sheet.cell(12, col_idx).value
    air_flow = sheet.cell(13, col_idx).value
    defrost_power = sheet.cell(14, col_idx).value
    pipe_dia = sheet.cell(15, col_idx).value
    noise = sheet.cell(16, col_idx).value
    weight = sheet.cell(17, col_idx).value
    series = sheet.cell(18, col_idx).value
    comment = sheet.cell(19, col_idx).value
    fin_spacing = sheet.cell(20, col_idx).value
    
    # 从fin_spacing中提取数字到fan_spacing_num
    fan_spacing_num = extract_fin_spacing_num(fin_spacing)
    
    model_sql = format_sql_value(model)
    heat_exchange_area_sql = format_sql_value(heat_exchange_area)
    tube_volumn_sql = format_sql_value(tube_volumn)
    air_flow_rate_sql = format_sql_value(air_flow_rate)
    total_fan_power_sql = format_sql_value(total_fan_power)
    total_fan_current_sql = format_sql_value(total_fan_current)
    air_flow_sql = format_sql_value(air_flow)
    defrost_power_sql = format_sql_value(defrost_power)
    pipe_dia_sql = format_sql_value(pipe_dia)
    noise_sql = format_sql_value(noise)
    weight_sql = format_sql_value(weight)
    fin_spacing_sql = format_sql_value(fin_spacing)
    fan_spacing_num_sql = format_sql_value(fan_spacing_num)
    series_sql = format_sql_value(series)
    comment_sql = format_sql_value(comment)
    
    insert_statement = (
        f"INSERT INTO cooler (model, heat_exchange_area, tube_volumn, air_flow_rate, "
        f"total_fan_power, total_fan_current, air_flow, defrost_power, "
        f"pipe_dia, noise, weight, fin_spacing, fan_spacing_num, series, comment, "
        f"create_time, update_time, is_deleted) "
        f"VALUES ({model_sql}, {heat_exchange_area_sql}, {tube_volumn_sql}, {air_flow_rate_sql}, "
        f"{total_fan_power_sql}, {total_fan_current_sql}, {air_flow_sql}, {defrost_power_sql}, "
        f"{pipe_dia_sql}, {noise_sql}, {weight_sql}, {fin_spacing_sql}, {fan_spacing_num_sql}, "
        f"{series_sql}, {comment_sql}, NOW(), NOW(), 0);"
    )
    
    return insert_statement


def generate_sql_from_excel(excel_file_path, cooling_capacity_output_path, cooler_output_path):
    """
    从Excel文件生成SQL INSERT语句
    
    Args:
        excel_file_path: Excel文件路径
        cooling_capacity_output_path: cooling_capacity表SQL输出路径
        cooler_output_path: cooler表SQL输出路径
    """
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook[workbook.sheetnames[0]]
    
    cooling_capacity_statements = []
    cooler_statements = []
    
    print(f"处理Excel文件: {excel_file_path}")
    print(f"总行数: {sheet.max_row}")
    print(f"总列数: {sheet.max_column}\n")
    
    for col_idx in range(2, sheet.max_column + 1):
        model = sheet.cell(1, col_idx).value
        
        if model is None or model == '':
            continue
        
        print(f"处理冷风机: {model}")
        
        refrigerant = sheet.cell(7, col_idx).value
        
        cooling_capacity_inserts = generate_cooling_capacity_insert(sheet, col_idx, refrigerant)
        cooling_capacity_statements.extend(cooling_capacity_inserts)
        
        cooler_insert = generate_cooler_insert(sheet, col_idx)
        if cooler_insert:
            cooler_statements.append(cooler_insert)
    
    with open(cooling_capacity_output_path, 'w', encoding='utf-8') as f:
        f.write(f"-- cooling_capacity 表数据插入语句\n")
        f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- 数据来源: {excel_file_path}\n\n")
        
        for statement in cooling_capacity_statements:
            f.write(statement + '\n')
    
    with open(cooler_output_path, 'w', encoding='utf-8') as f:
        f.write(f"-- cooler 表数据插入语句\n")
        f.write(f"-- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- 数据来源: {excel_file_path}\n\n")
        
        for statement in cooler_statements:
            f.write(statement + '\n')
    
    print(f"\n成功生成 {len(cooling_capacity_statements)} 条cooling_capacity INSERT语句")
    print(f"输出文件: {cooling_capacity_output_path}")
    print(f"\n成功生成 {len(cooler_statements)} 条cooler INSERT语句")
    print(f"输出文件: {cooler_output_path}")


def main():
    excel_file = "冷风机数据.xlsx"
    cooling_capacity_output_file = "app/sql/cooling_capacity.sql"
    cooler_output_file = "app/sql/cooler.sql"
    
    excel_file_abs = os.path.abspath(excel_file)
    cooling_capacity_output_abs = os.path.abspath(cooling_capacity_output_file)
    cooler_output_abs = os.path.abspath(cooler_output_file)
    
    try:
        generate_sql_from_excel(excel_file_abs, cooling_capacity_output_abs, cooler_output_abs)
        print("\n转换完成！")
        
    except FileNotFoundError as e:
        print(f"错误: 找不到文件 {excel_file_abs}")
        print(f"异常详情: {e}")
    except Exception as e:
        print(f"错误: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
